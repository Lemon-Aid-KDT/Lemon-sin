"""
v3.5: 마크다운 텍스트 → CSV / XLSX 변환 모듈

마크다운 내 테이블(|...|)과 Key: Value 패턴을 파싱하여
구조화된 표 형식으로 내보낸다.
"""
from __future__ import annotations

import csv
import io
import re


def _parse_markdown_tables(text: str) -> list[list[list[str]]]:
    """마크다운 테이블을 파싱하여 2차원 배열 리스트로 반환한다.

    Returns:
        list of tables, where each table is a list of rows (list of cell strings).
    """
    tables: list[list[list[str]]] = []
    current_table: list[list[str]] = []

    for line in text.split("\n"):
        stripped = line.strip()
        if stripped.startswith("|") and stripped.endswith("|"):
            # 구분선(|---|---|) 무시
            if re.match(r"^\|[\s\-:|]+\|$", stripped):
                continue
            cells = [c.strip() for c in stripped.split("|")[1:-1]]
            current_table.append(cells)
        else:
            if current_table:
                tables.append(current_table)
                current_table = []

    if current_table:
        tables.append(current_table)

    return tables


def _parse_key_value_pairs(text: str) -> list[tuple[str, str]]:
    """**Key**: Value 또는 Key: Value 패턴을 추출한다."""
    pairs: list[tuple[str, str]] = []
    pattern = re.compile(r"^\s*\**([^*:]+?)\**\s*[:：]\s*(.+)$", re.MULTILINE)
    for match in pattern.finditer(text):
        key = match.group(1).strip()
        value = match.group(2).strip()
        # 마크다운 볼드/이탈릭 제거
        value = re.sub(r"\*+", "", value)
        pairs.append((key, value))
    return pairs


def text_to_csv_bytes(text: str, title: str = "") -> bytes:
    """마크다운 텍스트를 CSV 바이트로 변환한다.

    우선순위: 테이블 > Key-Value > 줄 단위 fallback
    BOM(utf-8-sig) 포함하여 Excel에서 한글 정상 표시.
    """
    buf = io.StringIO()
    writer = csv.writer(buf)

    tables = _parse_markdown_tables(text)
    kv_pairs = _parse_key_value_pairs(text)

    if tables:
        # 테이블이 있으면 모든 테이블을 순서대로 기록
        for i, table in enumerate(tables):
            if i > 0:
                writer.writerow([])  # 테이블 간 빈 줄
            for row in table:
                writer.writerow(row)
    elif kv_pairs:
        # Key-Value 쌍을 2열 테이블로
        writer.writerow(["항목", "내용"])
        for key, value in kv_pairs:
            writer.writerow([key, value])
    else:
        # fallback: 줄 단위 출력
        if title:
            writer.writerow([title])
        for line in text.split("\n"):
            stripped = line.strip()
            if stripped:
                # 마크다운 헤더 기호 제거
                cleaned = re.sub(r"^#+\s*", "", stripped)
                writer.writerow([cleaned])

    return buf.getvalue().encode("utf-8-sig")


def text_to_xlsx_bytes(text: str, title: str = "") -> bytes:
    """마크다운 텍스트를 XLSX 바이트로 변환한다.

    openpyxl 사용, 헤더 볼드 + 열 폭 자동 조절.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
    except ImportError:
        # openpyxl 미설치 시 CSV fallback
        return text_to_csv_bytes(text, title)

    wb = Workbook()
    ws = wb.active
    # v3.5 fix: openpyxl 시트명에 금지 문자( / \ * ? : [ ] ) 제거 + 31자 제한
    _safe_title = re.sub(r"[/\\*?\[\]:]", "_", title or "Sheet1")
    ws.title = _safe_title[:31]

    tables = _parse_markdown_tables(text)
    kv_pairs = _parse_key_value_pairs(text)

    bold_font = Font(bold=True)
    wrap_align = Alignment(wrap_text=True, vertical="top")

    if tables:
        row_idx = 1
        for t_idx, table in enumerate(tables):
            if t_idx > 0:
                row_idx += 1  # 테이블 간 빈 행
            for r_idx, row in enumerate(table):
                for c_idx, cell in enumerate(row, start=1):
                    c = ws.cell(row=row_idx, column=c_idx, value=cell)
                    c.alignment = wrap_align
                    if r_idx == 0:
                        c.font = bold_font
                row_idx += 1
    elif kv_pairs:
        ws.cell(row=1, column=1, value="항목").font = bold_font
        ws.cell(row=1, column=2, value="내용").font = bold_font
        for i, (key, value) in enumerate(kv_pairs, start=2):
            ws.cell(row=i, column=1, value=key)
            ws.cell(row=i, column=2, value=value).alignment = wrap_align
    else:
        if title:
            ws.cell(row=1, column=1, value=title).font = bold_font
        start = 2 if title else 1
        for i, line in enumerate(text.split("\n")):
            stripped = line.strip()
            if stripped:
                cleaned = re.sub(r"^#+\s*", "", stripped)
                ws.cell(row=start + i, column=1, value=cleaned)

    # 열 폭 자동 조절
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
