"""Plan v1.0 §4.1 — 포맷별 reshape 레이어.

`/draft/export` 의 단일 진입점. 같은 마크다운 콘텐츠를 포맷 특성에 맞게 가공한다.

- txt:       UTF-8 평문
- docx:      python-docx (DocxExporter)
- pdf:       reportlab + Pretendard (PdfExporter — 한글 폰트 임베드)
- xlsx:      8D / ECN / 회의록은 다중 시트, 그 외는 단일 시트
- csv:       마크다운 표 우선, 없으면 key-value
- hwpx:      HwpxExporter (한글 정식)
- odt:       DOCX fallback (한 사이클 후 정식 ODT 로 교체 예정)
- clipboard: 프론트 처리 (백엔드 미사용)
"""
from __future__ import annotations

import io
import re
from typing import Literal

ShaperFormat = Literal["txt", "docx", "pdf", "xlsx", "csv", "hwpx", "odt"]

# ───────────────────────────────────────────────────────────
# 다중 시트 분기 — 8D Report / ECN / 회의록 등 구조화 문서
# ───────────────────────────────────────────────────────────

_MULTI_SHEET_DOC_TYPES: set[str] = {
    "8d_report", "report_8d",
    "ecn", "report_ecn",
    "meeting_min", "report_meeting", "meeting_minutes",
}


def _is_multi_sheet(doc_type: str) -> bool:
    return (doc_type or "").lower() in _MULTI_SHEET_DOC_TYPES


# ───────────────────────────────────────────────────────────
# 섹션 파서 — '## 제목' 단위로 나눠 시트 분리
# ───────────────────────────────────────────────────────────


def _split_sections(md: str) -> list[tuple[str, str]]:
    """마크다운을 H2(##) 섹션 단위로 분할. (title, body) 튜플 리스트."""
    sections: list[tuple[str, str]] = []
    cur_title = "Meta"
    buf: list[str] = []
    for line in md.split("\n"):
        m = re.match(r"^##\s+(.+)$", line.rstrip())
        if m:
            if buf:
                sections.append((cur_title, "\n".join(buf).strip()))
                buf = []
            cur_title = m.group(1).strip()
        else:
            buf.append(line)
    if buf:
        sections.append((cur_title, "\n".join(buf).strip()))
    # 빈 섹션 제거
    return [(t, b) for t, b in sections if b.strip()]


def _safe_sheet_name(name: str, fallback: str = "Sheet") -> str:
    """openpyxl 시트명 호환: 31자 제한 + 금지 문자 치환."""
    cleaned = re.sub(r"[/\\*?\[\]:]", "_", (name or fallback).strip())
    cleaned = cleaned[:31] or fallback
    return cleaned


# ───────────────────────────────────────────────────────────
# 메인 엔트리포인트
# ───────────────────────────────────────────────────────────


def shape_for_format(
    content_md: str,
    doc_type: str,
    fmt: ShaperFormat,
) -> bytes:
    """포맷별 reshape 단일 진입점.

    Args:
        content_md: LLM/Jinja2 가 생성한 마크다운 본문
        doc_type:   문서 유형 (8d_report / ecn / oem_email 등)
        fmt:        대상 포맷
    Returns:
        파일 바이트
    Raises:
        ValueError: 지원하지 않는 포맷
        ImportError: 필수 라이브러리 없음 (호출부에서 503 반환)
    """
    fmt = fmt.lower()  # type: ignore[assignment]

    if fmt == "txt":
        return content_md.encode("utf-8")

    if fmt == "docx":
        from features.draft.docx_exporter import DocxExporter
        return DocxExporter().export_bytes(content_md, doc_type)

    if fmt == "pdf":
        from features.draft.pdf_exporter import PdfExporter
        return PdfExporter().export_bytes(content_md)

    if fmt == "xlsx":
        if _is_multi_sheet(doc_type):
            return _shape_xlsx_multi_sheet(content_md, doc_type)
        from features.draft.tabular_exporter import text_to_xlsx_bytes
        return text_to_xlsx_bytes(content_md, title=doc_type or "draft")

    if fmt == "csv":
        from features.draft.tabular_exporter import text_to_csv_bytes
        return text_to_csv_bytes(content_md, title=doc_type or "draft")

    if fmt == "hwpx":
        from features.draft.hwpx_exporter import HwpxExporter
        return HwpxExporter().export_bytes(content_md, doc_type)

    if fmt == "odt":
        # 한 사이클 병행 후 정식 ODT 라이브러리(odfpy)로 교체 예정 — 현재는 DOCX fallback
        from features.draft.docx_exporter import DocxExporter
        return DocxExporter().export_bytes(content_md, doc_type)

    raise ValueError(f"지원하지 않는 포맷: {fmt}")


# ───────────────────────────────────────────────────────────
# XLSX 다중 시트 (8D / ECN / 회의록)
# ───────────────────────────────────────────────────────────


def _shape_xlsx_multi_sheet(md: str, doc_type: str) -> bytes:
    """8D Report 등 구조화 문서를 H2 섹션마다 시트 분리."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
    except ImportError:
        # openpyxl 없으면 단일 시트로 fallback
        from features.draft.tabular_exporter import text_to_xlsx_bytes
        return text_to_xlsx_bytes(md, title=doc_type or "draft")

    from features.draft.tabular_exporter import (
        _parse_key_value_pairs,
        _parse_markdown_tables,
    )

    sections = _split_sections(md)
    wb = Workbook()
    # 기본 시트 제거 (Sheet)
    if wb.worksheets:
        wb.remove(wb.active)

    bold = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")

    if not sections:
        sections = [("Document", md)]

    for title, body in sections:
        ws = wb.create_sheet(title=_safe_sheet_name(title))
        tables = _parse_markdown_tables(body)
        kv = _parse_key_value_pairs(body)

        if tables:
            row_idx = 1
            for t_idx, table in enumerate(tables):
                if t_idx > 0:
                    row_idx += 1
                for r_idx, row in enumerate(table):
                    for c_idx, cell in enumerate(row, start=1):
                        c = ws.cell(row=row_idx, column=c_idx, value=cell)
                        c.alignment = wrap
                        if r_idx == 0:
                            c.font = bold
                    row_idx += 1
        elif kv:
            ws.cell(row=1, column=1, value="항목").font = bold
            ws.cell(row=1, column=2, value="내용").font = bold
            for i, (k, v) in enumerate(kv, start=2):
                ws.cell(row=i, column=1, value=k)
                ws.cell(row=i, column=2, value=v).alignment = wrap
        else:
            # 본문을 줄 단위로
            for i, line in enumerate(body.split("\n"), start=1):
                stripped = line.strip()
                if stripped:
                    cleaned = re.sub(r"^#+\s*", "", stripped)
                    ws.cell(row=i, column=1, value=cleaned).alignment = wrap

        # 열 폭 자동 조정
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
